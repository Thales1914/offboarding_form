from django.contrib import admin
from django.urls import path, reverse
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied
from django.utils.html import format_html
from django.db.models import Count
from openpyxl import load_workbook
from .forms import DistratoForm
import os

from .models import Desligamento, Admissao, Distrato


class FiltroQtdDesligamentos(admin.SimpleListFilter):
    title = "Quantidade de desligamentos"
    parameter_name = "qtd_desligamentos"

    def lookups(self, request, model_admin):
        return [
            ("1", "1 desligamento"),
            ("5", "5 ou mais"),
            ("10", "10 ou mais"),
        ]

    def queryset(self, request, queryset):
        if self.value() == "1":
            return queryset.annotate(total=Count("criado_por")).filter(total=1)
        if self.value() == "5":
            return queryset.annotate(total=Count("criado_por")).filter(total__gte=5)
        if self.value() == "10":
            return queryset.annotate(total=Count("criado_por")).filter(total__gte=10)
        return queryset


@admin.register(Desligamento)
class DesligamentoAdmin(admin.ModelAdmin):
    list_display = (
        "nome",
        "codigo",
        "supervisor",
        "demissao",
        "area_atuacao",
        "criado_por",
        "qtd_desligamentos_colaborador",
    )
    search_fields = ("nome", "codigo", "area_atuacao")
    list_filter = ("area_atuacao", "demissao", "criado_por", FiltroQtdDesligamentos)

    fieldsets = (
        ('📌 Dados do Colaborador', {
            'fields': ('codigo', 'nome', 'contato', 'admissao', 'demissao', 'area_atuacao')
        }),
        ('📄 Motivo do desligamento', {
            'fields': ('motivo',)
        }),
        ('📦 Itens a devolver', {
            'fields': (
                'fardamento', 'chip_voz', 'chip_dados', 'tablet',
                'carregador_tablet', 'fone_tablet', 'catalogo',
                'bloco_pedido', 'carta_pedido_demissao', 'relatorio_inadimplencia'
            )
        }),
        ('🔎 Perguntas extras', {
            'fields': ('substituto', 'telemarketing', 'nova_contratacao')
        }),
        ('📝 Observações', {
            'fields': ('observacoes',)
        }),
    )

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.criado_por = request.user
        super().save_model(request, obj, form, change)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = qs.annotate(total_desligamentos=Count("criado_por"))
        if request.user.is_superuser or request.user.groups.filter(name="RH").exists():
            return qs
        return qs.filter(criado_por=request.user)

    def qtd_desligamentos_colaborador(self, obj):
        return Desligamento.objects.filter(criado_por=obj.criado_por).count()
    qtd_desligamentos_colaborador.short_description = "Qtd desligamentos"

    def has_export_permission(self, request):
        return request.user.is_superuser or request.user.groups.filter(name="RH").exists()

    def has_view_permission(self, request, obj=None):
        if obj is None:
            return True
        if request.user.is_superuser or request.user.groups.filter(name="RH").exists():
            return True
        return obj.criado_por == request.user

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<int:desligamento_id>/exportar_excel/',
                self.admin_site.admin_view(self.exportar_excel),
                name="rh_desligamento_exportar_excel_individual",
            ),
        ]
        return custom_urls + urls

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        export_url = reverse("admin:rh_desligamento_exportar_excel_individual", args=[object_id])
        extra_context['extra_buttons'] = format_html(
            f'<a class="button" style="margin-left:10px;" href="{export_url}">📤 Exportar Excel</a>'
        )
        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    def exportar_excel(self, request, desligamento_id):
        if not self.has_export_permission(request):
            raise PermissionDenied("Você não tem permissão para exportar este registro.")

        desligamento = Desligamento.objects.get(id=desligamento_id)

        modelo_path = os.path.join(os.path.dirname(__file__), "FORMULÁRIO DESLIGAMENTO RCA.xlsx")
        wb = load_workbook(modelo_path)
        ws = wb.active

        ws["B3"] = desligamento.supervisor or ""
        ws["E3"] = desligamento.demissao.strftime("%d/%m/%Y") if desligamento.demissao else ""
        ws["F2"] = desligamento.admissao.strftime("%d/%m/%Y") if desligamento.admissao else ""

        ws["A6"] = desligamento.codigo or ""
        ws["B6"] = desligamento.nome or ""
        ws["C6"] = desligamento.contato or ""
        ws["D6"] = desligamento.admissao.strftime("%d/%m/%Y") if desligamento.admissao else ""
        ws["E6"] = desligamento.demissao.strftime("%d/%m/%Y") if desligamento.demissao else ""
        ws["F6"] = desligamento.area_atuacao or ""

        ws["C7"] = desligamento.motivo or ""

        itens = [
            desligamento.fardamento,
            desligamento.chip_voz,
            desligamento.chip_dados,
            desligamento.tablet,
            desligamento.carregador_tablet,
            desligamento.fone_tablet,
            desligamento.catalogo,
            desligamento.bloco_pedido,
            desligamento.carta_pedido_demissao,
            desligamento.relatorio_inadimplencia,
        ]
        for i, valor in enumerate(itens, start=10):
            ws[f"A{i}"] = "SIM" if valor else "NÃO"

        ws["E22"] = "SIM" if desligamento.substituto else "NÃO"
        ws["E23"] = "SIM" if desligamento.telemarketing else "NÃO"
        ws["E24"] = "SIM" if desligamento.nova_contratacao else "NÃO"

        ws["C26"] = desligamento.observacoes or ""

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="desligamento_{desligamento.codigo}.xlsx"'
        wb.save(response)
        return response

@admin.register(Admissao)
class AdmissaoAdmin(admin.ModelAdmin):
    list_display = ("nome", "codigo", "supervisor", "data_admissao", "cargo", "criado_por")
    search_fields = ("nome", "codigo", "cpf", "cargo", "supervisor_responsavel")
    list_filter = ("cargo", "data_admissao", "criado_por")

    fieldsets = (
        ("📌 Dados Pessoais", {
            "fields": ("codigo", "nome", "nascimento", "naturalidade", "uf", "mae", "pai")
        }),
        ("🏠 Endereço", {
            "fields": ("endereco", "bairro", "cidade", "estado", "cep")
        }),
        ("📞 Contato", {
            "fields": ("fone", "email")
        }),
        ("📄 Documentos", {
            "fields": ("rg", "orgao_exp", "emissao", "cpf")
        }),
        ("🏦 Dados Bancários", {
            "fields": ("banco", "agencia", "conta", "operacao")
        }),
        ("💼 Condições de Admissão", {
            "fields": ("data_admissao", "cargo", "substituicao", "supervisor_responsavel", "coordenador")
        }),
        ("🔐 Conta Gov", {
            "fields": ("conta_gov", "senha_gov")
        }),
        ("📝 Observações", {
            "fields": ("observacoes",)
        }),
    )

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.criado_por = request.user
        super().save_model(request, obj, form, change)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<int:admissao_id>/exportar_excel/',
                self.admin_site.admin_view(self.exportar_excel),
                name="rh_admissao_exportar_excel_individual",
            ),
        ]
        return custom_urls + urls

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        export_url = reverse("admin:rh_admissao_exportar_excel_individual", args=[object_id])
        extra_context['extra_buttons'] = format_html(
            f'<a class="button" style="margin-left:10px;" href="{export_url}">📤 Exportar Excel</a>'
        )
        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    def exportar_excel(self, request, admissao_id):
        admissao = Admissao.objects.get(id=admissao_id)
        modelo_path = os.path.join(os.path.dirname(__file__), "FORMULÁRIO ADMISSAO RCA.xlsx")
        wb = load_workbook(modelo_path)
        ws = wb.active

        ws["G3"] = admissao.codigo or ""
        ws["B6"] = admissao.nome or ""
        ws["B7"] = admissao.nascimento.strftime("%d/%m/%Y") if admissao.nascimento else ""
        ws["E7"] = admissao.naturalidade or ""
        ws["B8"] = admissao.mae or ""
        ws["B9"] = admissao.pai or ""
        ws["B10"] = admissao.endereco or ""
        ws["B11"] = admissao.bairro or ""
        ws["F11"] = admissao.cep or ""
        ws["B12"] = admissao.cidade or ""
        ws["B13"] = admissao.fone or ""
        ws["E13"] = admissao.email or ""
        ws["B14"] = admissao.rg or ""
        ws["E14"] = admissao.orgao_exp or ""
        ws["G14"] = admissao.emissao.strftime("%d/%m/%Y") if admissao.emissao else ""
        ws["B15"] = admissao.cpf or ""
        ws["B16"] = "Caixa Econômica"  # fixo
        ws["E16"] = admissao.conta or ""
        ws["G16"] = admissao.operacao or ""
        ws["B18"] = admissao.data_admissao.strftime("%d/%m/%Y") if admissao.data_admissao else ""
        ws["D18"] = admissao.cargo or ""
        ws["F18"] = admissao.substituicao or ""
        ws["C19"] = admissao.supervisor_responsavel or ""
        ws["F19"] = admissao.coordenador or ""
        ws["B20"] = admissao.conta_gov or ""
        ws["D20"] = admissao.senha_gov or ""

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="admissao_{admissao.codigo}.xlsx"'
        wb.save(response)
        return response


@admin.register(Distrato)
class DistratoAdmin(admin.ModelAdmin):
    list_display = (
        "nome", "cpf", "data_admissao", "data_demissao",
        "total_geral", "total_ultimos_3_meses", "criado_por"
    )
    search_fields = ("nome", "cpf", "rg")
    list_filter = ("data_demissao", "criado_por")

    fieldsets = (
        ("📌 Dados do Representante", {
            "fields": ("nome", "cpf", "rg")
        }),
        ("📅 Datas", {
            "fields": ("data_admissao", "data_demissao")
        }),
        ("💰 Totais (somente campos amarelos)", {
            "fields": ("total_geral", "total_ultimos_3_meses")
        }),
        ("🏦 Dados Bancários", {
            "fields": ("banco", "agencia", "operacao", "conta_corrente", "titular", "telefone")
        }),
    )

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.criado_por = request.user
        super().save_model(request, obj, form, change)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<int:distrato_id>/exportar_excel/',
                self.admin_site.admin_view(self.exportar_excel),
                name="rh_distrato_exportar_excel_individual",
            ),
        ]
        return custom_urls + urls

    def change_view(self, request, object_id, form_url='', extra_context=None):
        extra_context = extra_context or {}
        export_url = reverse("admin:rh_distrato_exportar_excel_individual", args=[object_id])
        extra_context['extra_buttons'] = format_html(
             f'<a class="button" style="margin-left:10px;" href="{export_url}">📤 Exportar Distrato</a>'
        )
        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    def exportar_excel(self, request, distrato_id):
        distrato = Distrato.objects.get(id=distrato_id)
        modelo_path = os.path.join(os.path.dirname(__file__), "FORMULÁRIO DISTRATO RCA.xlsx")
        wb = load_workbook(modelo_path)
        ws = wb.active

        ws["B5"] = distrato.nome or ""
        ws["E5"] = distrato.cpf or ""
        ws["F5"] = distrato.rg or ""

        ws["B10"] = distrato.data_admissao.strftime("%d/%m/%Y") if distrato.data_admissao else ""
        ws["C10"] = distrato.data_demissao.strftime("%d/%m/%Y") if distrato.data_demissao else ""
        ws["B13"] = distrato.total_geral or 0
        ws["B16"] = distrato.total_ultimos_3_meses or 0

        ws["C23"] = distrato.banco or ""
        ws["C24"] = distrato.agencia or ""
        ws["C25"] = distrato.operacao or ""
        ws["C26"] = distrato.conta_corrente or ""
        ws["C27"] = distrato.titular or ""
        ws["C28"] = distrato.telefone or ""

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="distrato_{distrato.id}.xlsx"'
        wb.save(response)
        return response