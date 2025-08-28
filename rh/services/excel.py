import os
from openpyxl import load_workbook
from django.http import HttpResponse


def exportar_desligamento_excel(desligamento, modelo_path=None):
    """Exporta planilha de desligamento"""
    modelo_path = modelo_path or os.path.join(os.path.dirname(__file__), "..", "FORMULÁRIO DESLIGAMENTO RCA.xlsx")
    wb = load_workbook(modelo_path)
    ws = wb.active

    ws["B3"] = desligamento.supervisor or ""
    ws["E3"] = desligamento.demissao.strftime("%d/%m/%Y") if desligamento.demissao else ""

    ws["A6"] = desligamento.codigo or ""
    ws["B6"] = desligamento.nome or ""
    ws["C6"] = desligamento.contato or ""
    ws["D6"] = desligamento.admissao.strftime("%d/%m/%Y") if desligamento.admissao else ""
    ws["E6"] = desligamento.demissao.strftime("%d/%m/%Y") if desligamento.demissao else ""
    ws["F6"] = desligamento.area_atuacao or ""

    ws["C7"] = desligamento.motivo or ""

    itens = [
        desligamento.fardamento,
        desligamento.chip_voz,
        desligamento.chip_dados,    
        desligamento.tablet,
        desligamento.carregador_tablet,
        desligamento.fone_tablet,
        desligamento.catalogo,
        desligamento.bloco_pedido,
        desligamento.carta_pedido_demissao,
        desligamento.relatorio_inadimplencia,
    ]
    for i, valor in enumerate(itens, start=10):
        ws[f"A{i}"] = "SIM" if valor else "NÃO"

    ws["E22"] = "SIM" if desligamento.substituto else "NÃO"
    ws["E23"] = "SIM" if desligamento.telemarketing else "NÃO"
    ws["E24"] = "SIM" if desligamento.nova_contratacao else "NÃO"


    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="desligamento_{desligamento.codigo}.xlsx"'
    wb.save(response)
    return response


def exportar_admissao_excel(admissao, modelo_path=None):
    """Exporta planilha de admissão"""
    modelo_path = modelo_path or os.path.join(os.path.dirname(__file__), "..", "FORMULÁRIO ADMISSAO RCA.xlsx")
    wb = load_workbook(modelo_path)
    ws = wb.active

    ws["G3"] = admissao.codigo or ""
    ws["B6"] = admissao.nome or ""
    ws["B7"] = admissao.nascimento.strftime("%d/%m/%Y") if admissao.nascimento else ""
    ws["E7"] = admissao.naturalidade or ""
    ws["B8"] = admissao.mae or ""
    ws["B9"] = admissao.pai or ""
    ws["B10"] = admissao.endereco or ""
    ws["B11"] = admissao.bairro or ""
    ws["F11"] = admissao.cep or ""
    ws["B12"] = admissao.cidade or ""
    ws["B13"] = admissao.fone or ""
    ws["E13"] = admissao.email or ""
    ws["B14"] = admissao.rg or ""
    ws["E14"] = admissao.orgao_exp or ""
    ws["G14"] = admissao.emissao.strftime("%d/%m/%Y") if admissao.emissao else ""
    ws["B15"] = admissao.cpf or ""
    ws["B16"] = admissao.agencia or ""
    ws["E16"] = admissao.conta or ""
    ws["G16"] = admissao.operacao or ""
    ws["B18"] = admissao.data_admissao.strftime("%d/%m/%Y") if admissao.data_admissao else ""
    ws["D18"] = admissao.cargo or ""
    ws["F18"] = "Sim" if admissao.substituicao else "Não"
    ws["C19"] = admissao.supervisor_responsavel or ""
    ws["F19"] = admissao.coordenador or ""
    ws["B20"] = admissao.conta_gov or ""
    ws["D20"] = admissao.senha_gov or ""

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="admissao_{admissao.codigo}.xlsx"'
    wb.save(response)
    return response


def exportar_distrato_excel(distrato, modelo_path=None):
    """Exporta planilha de distrato"""
    modelo_path = modelo_path or os.path.join(os.path.dirname(__file__), "..", "FORMULÁRIO DISTRATO RCA.xlsx")
    wb = load_workbook(modelo_path)
    ws = wb.active

    ws["B5"] = distrato.nome or ""
    ws["E5"] = distrato.cpf or ""
    ws["F5"] = distrato.rg or ""
    ws["B10"] = distrato.data_admissao.strftime("%d/%m/%Y") if distrato.data_admissao else ""
    ws["C10"] = distrato.data_demissao.strftime("%d/%m/%Y") if distrato.data_demissao else ""
    ws["B13"] = distrato.total_geral or 0
    ws["B16"] = distrato.total_ultimos_3_meses or 0
    ws["C23"] = distrato.banco or ""
    ws["C24"] = distrato.agencia or ""
    ws["C25"] = distrato.operacao or ""
    ws["C26"] = distrato.conta_corrente or ""
    ws["C27"] = distrato.titular or ""
    ws["C28"] = distrato.telefone or ""

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="distrato_{distrato.id}.xlsx"'
    wb.save(response)
    return response
