from django.http import HttpResponse
from django.shortcuts import get_object_or_404
import openpyxl
from io import BytesIO
from .models import Desligamento
import os
from django.conf import settings

def exportar_excel_individual(request, pk):
    desligamento = get_object_or_404(Desligamento, pk=pk)

    # Caminho do modelo de planilha
    modelo_path = os.path.join(settings.BASE_DIR, "FORMULÁRIO DESLIGAMENTO RCA.xlsx")
    wb = openpyxl.load_workbook(modelo_path)
    ws = wb.active

    # 🔹 Preenche os campos
    ws["B3"] = desligamento.supervisor
    ws["E3"] = desligamento.demissao.strftime("%d/%m/%Y") if desligamento.demissao else ""
    ws["F2"] = desligamento.admissao.strftime("%d/%m/%Y") if desligamento.admissao else ""

    ws["A6"] = desligamento.codigo
    ws["B6"] = desligamento.nome
    ws["C6"] = desligamento.contato
    ws["D6"] = desligamento.admissao.strftime("%d/%m/%Y") if desligamento.admissao else ""
    ws["E6"] = desligamento.demissao.strftime("%d/%m/%Y") if desligamento.demissao else ""
    ws["F6"] = desligamento.area_atuacao

    # Motivo do desligamento → espalhando texto
    motivo = desligamento.motivo or ""
    partes = [motivo[i:i+20] for i in range(0, len(motivo), 20)]
    colunas = ["C", "D", "E", "F"]
    for i, parte in enumerate(partes[:4]):  # até 4 células
        ws[f"{colunas[i]}7"] = parte

    # Itens devolvidos (A10 até A19)
    itens = [
        ("Fardamento", desligamento.fardamento),
        ("Chip Voz", desligamento.chip_voz),
        ("Chip Dados", desligamento.chip_dados),
        ("Tablet", desligamento.tablet),
        ("Carregador Tablet", desligamento.carregador_tablet),
        ("Fone de ouvido Tablet", desligamento.fone_tablet),
        ("Catálogo", desligamento.catalogo),
        ("Bloco Pedido", desligamento.bloco_pedido),
        ("Carta Pedido Demissão", desligamento.carta_pedido_demissao),
        ("Relatório Inadimplência", desligamento.relatorio_inadimplencia),
    ]
    for i, (nome_item, valor) in enumerate(itens, start=10):
        ws[f"A{i}"] = f"{nome_item}: {'✅' if valor else '❌'}"

    # Perguntas extras
    ws["E22"] = "Substituto"
    ws["F22"] = "✅" if desligamento.substituto else "❌"

    ws["E23"] = "Telemarketing"
    ws["F23"] = "✅" if desligamento.telemarketing else "❌"

    ws["E24"] = "Nova Contratação"
    ws["F24"] = "✅" if desligamento.nova_contratacao else "❌"

    # 🔹 Exportar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="desligamento_{pk}.xlsx"'
    return response
