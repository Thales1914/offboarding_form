import openpyxl
from django.http import HttpResponse
from .models import Desligamento
from io import BytesIO


def exportar_ficha_excel(desligamento_id):
    desligamento = Desligamento.objects.get(pk=desligamento_id)

    # Abre o modelo base
    wb = openpyxl.load_workbook("FORMULÁRIO DESLIGAMENTO RCA.xlsx")
    ws = wb.active

    # ==== CAMPOS CABEÇALHO ====
    ws["B3"] = desligamento.supervisor
    ws["E3"] = desligamento.demissao.strftime("%d/%m/%Y") if desligamento.demissao else ""
    ws["F2"] = desligamento.admissao.strftime("%d/%m/%Y") if desligamento.admissao else ""

    # ==== DADOS PRINCIPAIS ====
    ws["A6"] = desligamento.codigo
    ws["B6"] = desligamento.nome
    ws["C6"] = desligamento.contato
    ws["D6"] = desligamento.admissao.strftime("%d/%m/%Y") if desligamento.admissao else ""
    ws["E6"] = desligamento.demissao.strftime("%d/%m/%Y") if desligamento.demissao else ""
    ws["F6"] = desligamento.area_atuacao

    # ==== MOTIVO DO DESLIGAMENTO ====
    ws["C7"] = desligamento.motivo  # pode ocupar C7, D7, E7, F7 se for mesclado no modelo

    # ==== ITENS DEVOLVIDOS (A10 até A19) ====
    itens = [
        desligamento.fardamento,
        desligamento.chip_voz,
        desligamento.chip_dados,
        desligamento.tablet,
        desligamento.carregador_tablet,
        desligamento.fone_tablet,
        desligamento.catalogo,
        desligamento.bloco_pedido,
        desligamento.carta_pedido_demissao,
        desligamento.relatorio_inadimplencia,
    ]
    for idx, valor in enumerate(itens, start=10):
        ws[f"A{idx}"] = "✅" if valor else "❌"

    # ==== CAMPOS EXTRAS (E22:F24) ====
    ws["E22"] = "Sim" if desligamento.substituto else "Não"
    ws["E23"] = "Sim" if desligamento.telemarketing else "Não"
    ws["E24"] = "Sim" if desligamento.nova_contratacao else "Não"

    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Resposta HTTP para download
    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="ficha_{desligamento.codigo}.xlsx"'
    return response
